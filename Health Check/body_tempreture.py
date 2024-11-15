import openpyxl as px
import random
from decimal import Decimal, ROUND_HALF_UP

this_month = input("What month do you want to output? - ")

wb1 = px.load_workbook('C:\\Users\\10x\\OneDrive\\デスクトップ\\Shinjo Tobatsu\\1.National Institute of Technology\\Health Check\\Data.xlsx')
sh1 = wb1['健康記録票' + str(this_month) + '月']

wb2 = px.load_workbook('C:\\Users\\10x\\OneDrive\\デスクトップ\\Shinjo Tobatsu\\1.National Institute of Technology\\Health Check\\Submission.xlsx')
sh2 = wb2.active

for date_row in range(3, 51):
    if int(date_row) % 3 == 0:
        for i in range(1, 16):
            body_tempreture = random.uniform(36.4, 36.9)
            body_tempreture = Decimal(str(body_tempreture)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
            sh2.cell(row = 4, column = int(date_row)).value = body_tempreture
    
            body_tempreture = random.uniform(36.4, 36.9)
            body_tempreture = Decimal(str(body_tempreture)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
            sh2.cell(row = 5, column = int(date_row)).value = float(body_tempreture)

            body_tempreture = random.uniform(36.4, 36.9)
            body_tempreture = Decimal(str(body_tempreture)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
            sh2.cell(row = 20, column = int(date_row)).value = float(body_tempreture)

            body_tempreture = random.uniform(36.4, 36.9)
            body_tempreture = Decimal(str(body_tempreture)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP)
            sh2.cell(row = 21, column = int(date_row)).value = float(body_tempreture)

for i in range(1,35):
    for j in range(1,51):
        copy = sh1.cell(row = i, column = j).value
        sh2.cell(row = i, column =j, value = copy)

sh2['AC1'] = 'E2119'
sh2['AO1'] = '新庄 都逸'

wb2.save('C:\\Users\\10x\\OneDrive\\デスクトップ\\Shinjo Tobatsu\\1.National Institute of Technology\\Health Check\\Submission.xlsx')

wb1.close()
wb2.close()